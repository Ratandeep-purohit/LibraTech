from flask import Flask, render_template, request, redirect, url_for, flash, session, jsonify, abort
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, login_user, logout_user, login_required, current_user, UserMixin
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps
from datetime import datetime, timedelta

import os
from werkzeug.utils import secure_filename
from config import Config
from models import db, User, Book, Category, Author, Issue, Fine, AuditLog, UserRole, FeeHeader, StudentFee, FeeCollection, FeeCollectionItem, BookRequest, Notification
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from flask import send_file
import io

app = Flask(__name__)
app.config.from_object(Config)

# File Upload configuration
UPLOAD_FOLDER = 'static/uploads/books'
USER_UPLOAD_FOLDER = 'static/uploads/users'
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif'}
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['USER_UPLOAD_FOLDER'] = USER_UPLOAD_FOLDER

# Ensure upload directories exist
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(USER_UPLOAD_FOLDER, exist_ok=True)

def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

db.init_app(app)

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

# --- HELPERS & DECORATORS ---

@app.context_processor
def inject_now():
    return {'now': datetime.utcnow()} # Useful for templates

class UserWrapper(UserMixin):
    def __init__(self, user_model):
        self.id = user_model.id
        self.user_model = user_model
        self.role = user_model.role.value
        self.full_name = user_model.full_name

    def get_id(self):
        return str(self.id)

@login_manager.user_loader
def load_user(user_id):
    u = User.query.get(int(user_id))
    if u:
        return UserWrapper(u)
    return None

def role_required(allowed_roles):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if not current_user.is_authenticated:
                return redirect(url_for('login'))
            if current_user.role not in allowed_roles:
                flash('You do not have permission to access this page.', 'danger')
                return redirect(url_for('dashboard'))
            return f(*args, **kwargs)
        return decorated_function
    return decorator

# --- AUTH ROUTES ---

@app.route('/')
def landing():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    return render_template('landing.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        # Can login with username or email
        identifier = request.form.get('identifier')
        password = request.form.get('password')
        
        user = User.query.filter((User.username == identifier) | (User.email == identifier)).first()
        
        if user and user.check_password(password):
            login_user(UserWrapper(user))
            # Log audit
            log = AuditLog(action='LOGIN', details=f'User {user.username} logged in', user_id=user.id, ip_address=request.remote_addr)
            db.session.add(log)
            db.session.commit()
            
            flash('Login successful!', 'success')
            return redirect(url_for('dashboard'))
        else:
            flash('Invalid credentials.', 'danger')
            
    return render_template('auth/login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('You have been logged out.', 'info')
    return redirect(url_for('login'))

@app.route('/register', methods=['GET', 'POST'])
def register():
    # Only allow registration if no admin exists (first run) or if admin is creating? 
    # Or public registration for students? Specification says "Student Registration".
    # Let's allow public student registration for now.
    if request.method == 'POST':
        username = request.form.get('username')
        email = request.form.get('email')
        password = request.form.get('password')
        full_name = request.form.get('full_name')
        
        if User.query.filter_by(username=username).first() or User.query.filter_by(email=email).first():
            flash('Username or Email already exists.', 'danger')
        else:
            # Profile Picture handling
            profile_picture = 'default_user.jpg'
            if 'profile_picture' in request.files:
                file = request.files['profile_picture']
                if file and allowed_file(file.filename):
                    filename = secure_filename(f"{username}_{file.filename}")
                    file.save(os.path.join(app.config['USER_UPLOAD_FOLDER'], filename))
                    profile_picture = filename

            new_user = User(
                username=username, 
                email=email, 
                full_name=full_name,
                role=UserRole.STUDENT,
                profile_picture=profile_picture
            )
            new_user.set_password(password)
            db.session.add(new_user)
            db.session.commit()
            
            flash('Registration successful! Please login.', 'success')
            return redirect(url_for('login'))
            
    return render_template('auth/register.html')

@app.route('/profile', methods=['GET', 'POST'])
@login_required
def profile():
    if request.method == 'POST':
        try:
            current_user.full_name = request.form.get('full_name')
            current_user.email = request.form.get('email')
            current_user.contact_number = request.form.get('contact_number')
            current_user.address = request.form.get('address')
            
            # Update Profile Picture
            if 'profile_picture' in request.files:
                file = request.files['profile_picture']
                if file and allowed_file(file.filename):
                    filename = secure_filename(f"{current_user.username}_{file.filename}")
                    file.save(os.path.join(app.config['USER_UPLOAD_FOLDER'], filename))
                    current_user.profile_picture = filename
                    
            db.session.commit()
            flash('Profile updated successfully!', 'success')
        except Exception as e:
            db.session.rollback()
            flash(f'Error updating profile: {str(e)}', 'danger')
            
    return render_template('profile.html')

# --- DASHBOARD & COMMON ROUTES ---

@app.route('/dashboard')
@login_required
def dashboard():
    if current_user.role == 'admin':
        return redirect(url_for('admin_dashboard'))
    elif current_user.role == 'librarian':
        return redirect(url_for('librarian_dashboard'))
    elif current_user.role == 'student':
        return redirect(url_for('student_dashboard'))
    return "Unknown Role"

@app.route('/admin/dashboard')
@login_required
@role_required(['admin'])
def admin_dashboard():
    # Stats
    total_books = Book.query.count()
    total_students = User.query.filter_by(role=UserRole.STUDENT).count()
    issued_books = Issue.query.filter_by(status='issued').count()
    total_fines = db.session.query(db.func.sum(Fine.amount)).scalar() or 0
    
    # Fees Stats
    total_assigned = db.session.query(db.func.sum(StudentFee.amount)).scalar() or 0
    total_collected_items = db.session.query(db.func.sum(FeeCollectionItem.amount_collected)).scalar() or 0
    total_fees_collected = db.session.query(db.func.sum(FeeCollection.total_amount)).scalar() or 0
    total_fees_due = total_assigned - total_collected_items
    
    return render_template('admin/dashboard.html', 
                           total_books=total_books, 
                           total_students=total_students,
                           issued_books=issued_books,
                           total_fines=total_fines,
                           total_fees_collected=total_fees_collected,
                           total_fees_due=total_fees_due)

@app.route('/librarian/dashboard')
@login_required
@role_required(['librarian', 'admin'])
def librarian_dashboard():
    due_today = Issue.query.filter(db.func.date(Issue.due_date) == datetime.utcnow().date(), Issue.status == 'issued').count()
    pending_returns = Issue.query.filter(Issue.due_date < datetime.utcnow(), Issue.status == 'issued').count()
    return render_template('librarian/dashboard.html', due_today=due_today, pending_returns=pending_returns)

@app.route('/admin/add_student', methods=['GET', 'POST'])
@login_required
@role_required(['admin'])
def add_student():
    if request.method == 'POST':
        try:
            username = request.form.get('username')
            email = request.form.get('email')
            password = request.form.get('password')
            full_name = request.form.get('full_name')
            
            # New Fields
            contact_number = request.form.get('contact_number')
            address = request.form.get('address')
            enrollment_number = request.form.get('enrollment_number')
            semester = request.form.get('semester')
            program = request.form.get('program')
            # Joining date
            joining_date_str = request.form.get('joining_date')
            joining_date = datetime.strptime(joining_date_str, '%Y-%m-%d') if joining_date_str else datetime.utcnow()

            # Handling Profile Picture
            profile_picture = 'default_user.jpg'
            if 'profile_picture' in request.files:
                file = request.files['profile_picture']
                if file and allowed_file(file.filename):
                    filename = secure_filename(f"{username}_{file.filename}")
                    file.save(os.path.join(app.config['USER_UPLOAD_FOLDER'], filename))
                    profile_picture = filename

            if User.query.filter((User.username == username) | (User.email == email)).first():
                flash('User already exists!', 'danger')
            else:
                new_user = User(
                    username=username,
                    email=email,
                    full_name=full_name,
                    role=UserRole.STUDENT,
                    contact_number=contact_number,
                    address=address,
                    enrollment_number=enrollment_number,
                    semester=semester,
                    program=program,
                    joining_date=joining_date,
                    profile_picture=profile_picture
                )
                new_user.set_password(password)
                db.session.add(new_user)
                db.session.commit()
                flash(f'Student {full_name} added successfully!', 'success')
                return redirect(url_for('student_list'))
        except Exception as e:
            flash(f'Error adding student: {str(e)}', 'danger')
            
    return render_template('admin/add_student.html')

@app.route('/admin/edit_student/<int:user_id>', methods=['GET', 'POST'])
@login_required
@role_required(['admin'])
def edit_student(user_id):
    student = User.query.get_or_404(user_id)
    if student.role != UserRole.STUDENT:
        flash('Cannot edit non-student users here.', 'warning')
        return redirect(url_for('student_list'))

    if request.method == 'POST':
        try:
            student.username = request.form.get('username')
            student.email = request.form.get('email')
            student.full_name = request.form.get('full_name')
            
            # Update New Fields
            student.contact_number = request.form.get('contact_number')
            student.address = request.form.get('address')
            student.enrollment_number = request.form.get('enrollment_number')
            student.semester = request.form.get('semester')
            student.program = request.form.get('program')
            
            # Update Profile Picture
            if 'profile_picture' in request.files:
                file = request.files['profile_picture']
                if file and allowed_file(file.filename):
                    filename = secure_filename(f"{student.username}_{file.filename}")
                    file.save(os.path.join(app.config['USER_UPLOAD_FOLDER'], filename))
                    student.profile_picture = filename
            
            # Update password only if provided
            new_password = request.form.get('password')
            if new_password:
                student.set_password(new_password)
            
            db.session.commit()
            flash(f'Student {student.full_name} updated successfully!', 'success')
            return redirect(url_for('student_list'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error updating student: {str(e)}', 'danger')
            
    return render_template('admin/edit_student.html', student=student)

@app.route('/students')
@login_required
@role_required(['admin', 'librarian'])
def student_list():
    page = request.args.get('page', 1, type=int)
    search = request.args.get('q', '')
    
    query = User.query.filter_by(role=UserRole.STUDENT)
    if search:
        query = query.filter(User.full_name.ilike(f'%{search}%') | User.username.ilike(f'%{search}%'))
        
    students = query.paginate(page=page, per_page=10)
    return render_template('admin/student_list.html', students=students, search=search)

@app.route('/admin/import_students_page')
@login_required
@role_required(['admin'])
def import_students_page():
    return render_template('admin/student_bulk_import.html')

@app.route('/admin/download_student_sample')
@login_required
@role_required(['admin'])
def download_student_sample():
    import pandas as pd
    import io
    from flask import send_file

    columns = ['Full Name', 'Username', 'Email', 'Password', 'Contact Number', 'Enrollment Number', 'Program', 'Semester', 'Address']
    
    sample_data = [
        {
            'Full Name': 'John Doe',
            'Username': 'johndoe2026',
            'Email': 'john@example.com',
            'Password': 'password123',
            'Contact Number': '1234567890',
            'Enrollment Number': 'ENR001',
            'Program': 'BCA',
            'Semester': '1',
            'Address': '123 Main St, City'
        }
    ]
    
    df = pd.DataFrame(sample_data, columns=columns)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False, sheet_name='Students')
    
    output.seek(0)
    return send_file(output, 
                     download_name="Sample_Student_Import.xlsx", 
                     as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/admin/import_students', methods=['POST'])
@login_required
@role_required(['admin'])
def import_students():
    import pandas as pd
    file = request.files.get('file')
    if not file or not file.filename.endswith(('.xlsx', '.xls')):
        flash('Please upload a valid Excel file.', 'danger')
        return redirect(url_for('import_students_page'))
    
    try:
        df = pd.read_excel(file)
        count = 0
        for _, row in df.iterrows():
            full_name = str(row.get('Full Name', '')).strip()
            username = str(row.get('Username', '')).strip()
            email = str(row.get('Email', '')).strip()
            password = str(row.get('Password', 'student123')).strip()
            
            if full_name == 'nan' or not full_name or username == 'nan' or not username:
                continue
                
            if User.query.filter((User.username == username) | (User.email == email)).first():
                continue
                
            new_user = User(
                username=username,
                email=email,
                full_name=full_name,
                role=UserRole.STUDENT,
                contact_number=str(row.get('Contact Number', '')),
                enrollment_number=str(row.get('Enrollment Number', '')),
                program=str(row.get('Program', '')),
                semester=str(row.get('Semester', '')),
                address=str(row.get('Address', ''))
            )
            if password == 'nan' or not password:
                password = 'student123'
            new_user.set_password(password)
            db.session.add(new_user)
            count += 1
            
        db.session.commit()
        flash(f'Successfully imported {count} students!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error importing students: {str(e)}', 'danger')
        
    return redirect(url_for('student_list'))

@app.route('/admin/delete_student/<int:user_id>', methods=['POST'])
@login_required
@role_required(['admin'])
def delete_student(user_id):
    student = User.query.get_or_404(user_id)
    if student.role != UserRole.STUDENT:
        flash('Cannot delete non-student users.', 'danger')
        return redirect(url_for('student_list'))
    
    # Check if student has pending issues or unpaid fines before deleting?
    # For now, let's assume we can Soft Delete or just Hard Delete if no constraints
    # SRS mentions "Soft Delete" for books, let's try Hard Delete for users first or check constraints
    # If constraints (foreign keys) exist, we might fail.
    
    try:
        db.session.delete(student)
        db.session.commit()
        flash(f'Student {student.full_name} deleted successfully.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting student: {str(e)}', 'danger')
        
    return redirect(url_for('student_list'))

@app.route('/admin/export_students')
@login_required
@role_required(['admin', 'librarian'])
def export_students():
    students = User.query.filter_by(role=UserRole.STUDENT).all()
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Students Report"
    
    # Define styles
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    alt_fill_1 = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
    alt_fill_2 = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = [
        "S.No", "Full Name", "Username", "Email", "Enrollment No", 
        "Program", "Semester", "Contact Number", "Joining Date", "Address"
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Data rows
    for row_num, student in enumerate(students, 2):
        row_data = [
            row_num - 1,
            student.full_name,
            student.username,
            student.email,
            student.enrollment_number or 'N/A',
            student.program or 'N/A',
            student.semester or 'N/A',
            student.contact_number or 'N/A',
            student.joining_date.strftime('%d-%b-%Y') if student.joining_date else 'N/A',
            student.address or 'N/A'
        ]
        
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = border
            cell.alignment = Alignment(horizontal='left' if col_num in [2, 3, 4, 10] else 'center', vertical='center')
            
            # Alternating row colors
            cell.fill = alt_fill_1 if row_num % 2 == 0 else alt_fill_2
    
    # Adjust column widths
    column_widths = [8, 25, 15, 30, 15, 20, 10, 15, 15, 35]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + col_num)].width = width
    
    # Add summary
    summary_row = len(students) + 3
    ws.cell(row=summary_row, column=1).value = "SUMMARY"
    ws.cell(row=summary_row, column=1).font = Font(bold=True, size=12)
    
    ws.cell(row=summary_row + 1, column=1).value = "Total Students:"
    ws.cell(row=summary_row + 1, column=2).value = len(students)
    ws.cell(row=summary_row + 1, column=1).font = Font(bold=True)
    
    # Program-wise count
    programs = {}
    for student in students:
        prog = student.program or 'Unknown'
        programs[prog] = programs.get(prog, 0) + 1
    
    ws.cell(row=summary_row + 3, column=1).value = "Program-wise Distribution:"
    ws.cell(row=summary_row + 3, column=1).font = Font(bold=True)
    
    for idx, (prog, count) in enumerate(programs.items(), 1):
        ws.cell(row=summary_row + 3 + idx, column=1).value = prog
        ws.cell(row=summary_row + 3 + idx, column=2).value = count
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"Students_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@app.route('/student/dashboard')
@login_required
@role_required(['student'])
def student_dashboard():
    # Only show active issues on dashboard
    my_issues_count = Issue.query.filter_by(user_id=current_user.id, status='issued').count()
    my_fines = Issue.query.join(Fine).filter(Issue.user_id == current_user.id, Fine.paid == False).with_entities(db.func.sum(Fine.amount)).scalar() or 0
    # Fetch active issues for the table
    current_issues = Issue.query.filter_by(user_id=current_user.id, status='issued').all()
    return render_template('student/dashboard.html', my_issues=my_issues_count, my_fines=my_fines, current_issues=current_issues)

@app.route('/student/history')
@login_required
@role_required(['student'])
def student_history():
    # Show all history for the student
    history = Issue.query.filter_by(user_id=current_user.id).order_by(Issue.issue_date.desc()).all()
    return render_template('student/history.html', history=history)

# --- BOOK MANAGEMENT (Admin/Librarian) ---

@app.route('/books')
@login_required
def book_list():
    page = request.args.get('page', 1, type=int)
    search = request.args.get('q', '')
    
    query = Book.query.filter(Book.is_deleted == False)
    if search:
        query = query.filter(Book.title.ilike(f'%{search}%') | Book.isbn.ilike(f'%{search}%'))
        
    books = query.paginate(page=page, per_page=10)
    return render_template('librarian/book_list.html', books=books, search=search)

@app.route('/books/add', methods=['GET', 'POST'])
@login_required
@role_required(['admin', 'librarian'])
def add_book():
    if request.method == 'POST':
        try:
            # Handle Image Upload
            cover_image = 'default_book.jpg'
            if 'book_image' in request.files:
                file = request.files['book_image']
                if file and allowed_file(file.filename):
                    filename = secure_filename(file.filename)
                    # Unique filename
                    unique_filename = f"{datetime.utcnow().strftime('%Y%m%d%H%M%S')}_{filename}"
                    file.save(os.path.join(app.config['UPLOAD_FOLDER'], unique_filename))
                    cover_image = unique_filename

            new_book = Book(
                title=request.form['title'],
                isbn=request.form['isbn'],
                category_id=request.form['category_id'],
                author_id=request.form['author_id'],
                publication_year=request.form['year'],
                publisher=request.form['publisher'],
                rack_number=request.form['rack'],
                total_copies=request.form['copies'],
                available_copies=request.form['copies'],
                cover_image=cover_image
            )
            db.session.add(new_book)
            db.session.commit()
            flash('Book added successfully!', 'success')
            return redirect(url_for('book_list'))
        except Exception as e:
            flash(f'Error adding book: {str(e)}', 'danger')
            
    categories = Category.query.all()
    authors = Author.query.all()
    return render_template('librarian/book_form.html', categories=categories, authors=authors)

@app.route('/books/edit/<int:book_id>', methods=['GET', 'POST'])
@login_required
@role_required(['admin', 'librarian'])
def edit_book(book_id):
    book = Book.query.get_or_404(book_id)
    if request.method == 'POST':
        try:
            # Handle Image Upload
            if 'book_image' in request.files:
                file = request.files['book_image']
                if file and allowed_file(file.filename):
                    filename = secure_filename(file.filename)
                    unique_filename = f"{datetime.utcnow().strftime('%Y%m%d%H%M%S')}_{filename}"
                    file.save(os.path.join(app.config['UPLOAD_FOLDER'], unique_filename))
                    book.cover_image = unique_filename

            book.title = request.form['title']
            book.isbn = request.form['isbn']
            book.category_id = request.form['category_id']
            book.author_id = request.form['author_id']
            book.publication_year = request.form['year']
            book.publisher = request.form['publisher']
            book.rack_number = request.form['rack']
            
            # Simple stock update: if total copies increase, available also increase
            old_total = book.total_copies
            new_total = int(request.form['copies'])
            diff = new_total - old_total
            book.total_copies = new_total
            book.available_copies += diff

            db.session.commit()
            flash('Book details updated!', 'success')
            return redirect(url_for('book_list'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error updating book: {str(e)}', 'danger')

    categories = Category.query.all()
    authors = Author.query.all()
    return render_template('librarian/book_form.html', book=book, categories=categories, authors=authors)

@app.route('/books/delete/<int:book_id>', methods=['POST'])
@login_required
@role_required(['admin', 'librarian'])
def delete_book(book_id):
    book = Book.query.get_or_404(book_id)
    # Check if currently issued?
    issued = Issue.query.filter_by(book_id=book.id, status='issued').first()
    if issued:
        flash('Cannot delete book that is currently issued.', 'warning')
        return redirect(url_for('book_list'))
    
    book.is_deleted = True
    db.session.commit()
    flash('Book removed from catalog.', 'success')
    return redirect(url_for('book_list'))

@app.route('/books/export')
@login_required
@role_required(['admin', 'librarian'])
def export_books():
    books = Book.query.filter(Book.is_deleted == False).all()
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Books Catalog"
    
    # Define styles
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    available_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
    unavailable_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    alt_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = [
        "S.No", "Title", "ISBN", "Author", "Category", "Publisher", 
        "Year", "Rack No", "Total Copies", "Available", "Status"
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Data rows
    for row_num, book in enumerate(books, 2):
        is_available = book.available_copies > 0
        
        row_data = [
            row_num - 1,
            book.title,
            book.isbn,
            book.author.name,
            book.category.name,
            book.publisher or 'N/A',
            book.publication_year or 'N/A',
            book.rack_number or 'N/A',
            book.total_copies,
            book.available_copies,
            "Available" if is_available else "Out of Stock"
        ]
        
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = border
            cell.alignment = Alignment(horizontal='left' if col_num in [2, 4, 5, 6] else 'center', vertical='center')
            
            # Color coding based on availability
            if is_available:
                cell.fill = available_fill
            else:
                cell.fill = unavailable_fill
    
    # Adjust column widths
    column_widths = [8, 40, 15, 25, 20, 25, 8, 10, 12, 12, 15]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + col_num)].width = width
    
    # Add summary
    summary_row = len(books) + 3
    ws.cell(row=summary_row, column=1).value = "SUMMARY"
    ws.cell(row=summary_row, column=1).font = Font(bold=True, size=12)
    
    total_books = len(books)
    available_books = sum(1 for b in books if b.available_copies > 0)
    out_of_stock = total_books - available_books
    total_copies = sum(b.total_copies for b in books)
    available_copies = sum(b.available_copies for b in books)
    
    ws.cell(row=summary_row + 1, column=1).value = "Total Books:"
    ws.cell(row=summary_row + 1, column=2).value = total_books
    ws.cell(row=summary_row + 1, column=1).font = Font(bold=True)
    
    ws.cell(row=summary_row + 2, column=1).value = "Available Books:"
    ws.cell(row=summary_row + 2, column=2).value = available_books
    ws.cell(row=summary_row + 2, column=1).font = Font(bold=True)
    ws.cell(row=summary_row + 2, column=2).fill = available_fill
    
    ws.cell(row=summary_row + 3, column=1).value = "Out of Stock:"
    ws.cell(row=summary_row + 3, column=2).value = out_of_stock
    ws.cell(row=summary_row + 3, column=1).font = Font(bold=True)
    ws.cell(row=summary_row + 3, column=2).fill = unavailable_fill
    
    ws.cell(row=summary_row + 5, column=1).value = "Total Copies:"
    ws.cell(row=summary_row + 5, column=2).value = total_copies
    ws.cell(row=summary_row + 5, column=1).font = Font(bold=True)
    
    ws.cell(row=summary_row + 6, column=1).value = "Available Copies:"
    ws.cell(row=summary_row + 6, column=2).value = available_copies
    ws.cell(row=summary_row + 6, column=1).font = Font(bold=True)
    
    # Category-wise distribution
    categories = {}
    for book in books:
        cat = book.category.name
        categories[cat] = categories.get(cat, 0) + 1
    
    ws.cell(row=summary_row + 8, column=1).value = "Category-wise Distribution:"
    ws.cell(row=summary_row + 8, column=1).font = Font(bold=True)
    
    for idx, (cat, count) in enumerate(categories.items(), 1):
        ws.cell(row=summary_row + 8 + idx, column=1).value = cat
        ws.cell(row=summary_row + 8 + idx, column=2).value = count
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"Books_Catalog_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@app.route('/books/import_page')
@login_required
@role_required(['admin', 'librarian'])
def import_page():
    return render_template('librarian/bulk_import.html')

@app.route('/books/download_sample')
@login_required
@role_required(['admin', 'librarian'])
def download_sample_excel():
    import pandas as pd
    import io
    from flask import send_file

    # Define the exact columns needed for the catalog
    columns = ['Title', 'ISBN', 'Category', 'Author', 'Year', 'Publisher', 'Rack', 'Total Copies']
    
    # Create sample data
    sample_data = [
        {
            'Title': 'The Great Gatsby',
            'ISBN': '9780743273565',
            'Category': 'Fiction',
            'Author': 'F. Scott Fitzgerald',
            'Year': 1925,
            'Publisher': 'Scribner',
            'Rack': 'A-101',
            'Total Copies': 5
        },
        {
            'Title': 'Clean Code',
            'ISBN': '9780132350884',
            'Category': 'Programming',
            'Author': 'Robert C. Martin',
            'Year': 2008,
            'Publisher': 'Prentice Hall',
            'Rack': 'B-205',
            'Total Copies': 3
        }
    ]
    
    df = pd.DataFrame(sample_data, columns=columns)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False, sheet_name='SampleData')
    
    output.seek(0)
    return send_file(output, 
                     download_name="Sample_Book_Import.xlsx", 
                     as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/books/import', methods=['POST'])
@login_required
@role_required(['admin', 'librarian'])
def import_books():
    import pandas as pd
    file = request.files.get('file')
    if not file or not file.filename.endswith(('.xlsx', '.xls')):
        flash('Please upload a valid Excel file.', 'danger')
        return redirect(url_for('import_page'))
    
    try:
        df = pd.read_excel(file)
        # Standard columns: Title, ISBN, Category, Author, Year, Publisher, Rack, Total Copies
        count = 0
        for _, row in df.iterrows():
            title = str(row.get('Title', '')).strip()
            isbn = str(row.get('ISBN', '')).strip()
            cat_name = str(row.get('Category', 'General')).strip()
            auth_name = str(row.get('Author', 'Unknown')).strip()
            
            # Handle NaN/Empty
            if title == 'nan' or not title or isbn == 'nan' or not isbn:
                continue
            
            # Check duplication
            if Book.query.filter_by(isbn=isbn).first():
                continue
            
            # Find/Create Category
            cat = Category.query.filter_by(name=cat_name).first()
            if not cat:
                cat = Category(name=cat_name)
                db.session.add(cat)
                db.session.flush()
                
            # Find/Create Author
            auth = Author.query.filter_by(name=auth_name).first()
            if not auth:
                auth = Author(name=auth_name)
                db.session.add(auth)
                db.session.flush()
                
            total_copies = int(row.get('Total Copies', 1)) if not pd.isna(row.get('Total Copies')) else 1

            new_book = Book(
                title=title,
                isbn=isbn,
                category_id=cat.id,
                author_id=auth.id,
                publication_year=int(row.get('Year', 2024)) if not pd.isna(row.get('Year')) else 2024,
                publisher=str(row.get('Publisher', '')) if not pd.isna(row.get('Publisher')) else '',
                rack_number=str(row.get('Rack', '')) if not pd.isna(row.get('Rack')) else '',
                total_copies=total_copies,
                available_copies=total_copies
            )
            db.session.add(new_book)
            count += 1
            
        db.session.commit()
        flash(f'Successfully imported {count} books from Excel!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error importing books: {str(e)}', 'danger')
        
    return redirect(url_for('book_list'))

# --- ISSUE & RETURN SYSTEM ---

@app.route('/issue/new', methods=['GET', 'POST'])
@login_required
@role_required(['admin', 'librarian'])
def issue_book():
    if request.method == 'POST':
        student_username = request.form.get('student_username')
        isbn = request.form.get('isbn')
        
        student = User.query.filter_by(username=student_username, role=UserRole.STUDENT).first()
        book = Book.query.filter_by(isbn=isbn).first()
        
        if not student:
            flash('Student not found.', 'danger')
            return redirect(url_for('issue_book'))
        if not book:
            flash('Book not found.', 'danger')
            return redirect(url_for('issue_book'))
        if book.available_copies < 1:
            flash('Book not available.', 'warning')
            return redirect(url_for('issue_book'))
            
        # Create issue
        new_issue = Issue(
            user_id=student.id,
            book_id=book.id,
            due_date=datetime.utcnow() + timedelta(days=14), # 2 weeks default
            status='issued'
        )
        book.available_copies -= 1
        db.session.add(new_issue)
        db.session.commit()
        
        flash(f'Book issued to {student.full_name}. Due date: {new_issue.due_date.date()}', 'success')
        return redirect(url_for('librarian_dashboard'))
        
    return render_template('librarian/issue_book.html')

@app.route('/issued-books')
@login_required
@role_required(['admin', 'librarian'])
def issued_books():
    page = request.args.get('page', 1, type=int)
    search = request.args.get('q', '')
    
    query = Issue.query.filter_by(status='issued')
    
    if search:
        query = query.join(User).join(Book).filter(
            db.or_(
                User.full_name.ilike(f'%{search}%'),
                User.username.ilike(f'%{search}%'),
                Book.title.ilike(f'%{search}%'),
                Book.isbn.ilike(f'%{search}%')
            )
        )
    
    issues = query.order_by(Issue.issue_date.desc()).paginate(page=page, per_page=20, error_out=False)
    
    return render_template('admin/issued_books.html', issues=issues, search=search, now=datetime.utcnow())

@app.route('/issued-books/export')
@login_required
@role_required(['admin', 'librarian'])
def export_issued_books():
    # Get all currently issued books
    issues = Issue.query.filter_by(status='issued').order_by(Issue.issue_date.desc()).all()
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Issued Books Report"
    
    # Define styles
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    overdue_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    active_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = [
        "S.No", "Book Title", "ISBN", "Author", "Student Name", "Username", 
        "Enrollment No", "Issue Date", "Days Issued", "Due Date", "Days Until Due", "Status"
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Data rows
    now = datetime.utcnow()
    for row_num, issue in enumerate(issues, 2):
        days_issued = (now - issue.issue_date).days
        days_until_due = (issue.due_date - now).days
        is_overdue = days_until_due < 0
        
        row_data = [
            row_num - 1,
            issue.book.title,
            issue.book.isbn,
            issue.book.author.name,
            issue.student.full_name,
            issue.student.username,
            issue.student.enrollment_number or 'N/A',
            issue.issue_date.strftime('%d-%b-%Y %I:%M %p'),
            days_issued,
            issue.due_date.strftime('%d-%b-%Y'),
            days_until_due if not is_overdue else f"-{abs(days_until_due)} (Overdue)",
            "OVERDUE" if is_overdue else "ACTIVE"
        ]
        
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = border
            cell.alignment = Alignment(horizontal='left' if col_num in [2, 4, 5, 6] else 'center', vertical='center')
            
            # Apply status-based coloring
            if is_overdue:
                cell.fill = overdue_fill
            else:
                cell.fill = active_fill
    
    # Adjust column widths
    column_widths = [8, 35, 15, 25, 25, 15, 15, 22, 12, 15, 18, 12]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + col_num)].width = width
    
    # Add summary at the bottom
    summary_row = len(issues) + 3
    ws.cell(row=summary_row, column=1).value = "SUMMARY"
    ws.cell(row=summary_row, column=1).font = Font(bold=True, size=12)
    
    total_issued = len(issues)
    overdue_count = sum(1 for issue in issues if (issue.due_date - now).days < 0)
    active_count = total_issued - overdue_count
    
    ws.cell(row=summary_row + 1, column=1).value = "Total Issued:"
    ws.cell(row=summary_row + 1, column=2).value = total_issued
    ws.cell(row=summary_row + 1, column=1).font = Font(bold=True)
    
    ws.cell(row=summary_row + 2, column=1).value = "Overdue:"
    ws.cell(row=summary_row + 2, column=2).value = overdue_count
    ws.cell(row=summary_row + 2, column=1).font = Font(bold=True)
    ws.cell(row=summary_row + 2, column=2).fill = overdue_fill
    
    ws.cell(row=summary_row + 3, column=1).value = "Active:"
    ws.cell(row=summary_row + 3, column=2).value = active_count
    ws.cell(row=summary_row + 3, column=1).font = Font(bold=True)
    ws.cell(row=summary_row + 3, column=2).fill = active_fill
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"Issued_Books_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )



@app.route('/return/<int:issue_id>', methods=['POST'])
@login_required
@role_required(['admin', 'librarian'])
def return_book_process(issue_id):
    issue = Issue.query.get_or_404(issue_id)
    if issue.status != 'issued':
        flash('Book already returned.', 'warning')
        return redirect(url_for('librarian_dashboard'))
        
    issue.return_date = datetime.utcnow()
    issue.status = 'returned'
    issue.book.available_copies += 1
    
    # Calculate fine logic (simplified)
    # If returned after due date
    if issue.return_date > issue.due_date:
        overdue_days = (issue.return_date - issue.due_date).days
        fine_amount = overdue_days * 5.0 # 5 currency units per day
        if fine_amount > 0:
            fine = Fine(amount=fine_amount, issue_id=issue.id)
            db.session.add(fine)
            flash(f'Book returned late! Fine: {fine_amount}', 'warning')
        else:
             flash('Book returned successfully.', 'success')
    else:
        flash('Book returned successfully.', 'success')
        
    db.session.commit()
    return redirect(url_for('librarian_dashboard'))

# --- ANALYTICS & REPORTS --- 

@app.route('/admin/analytics')
@login_required
@role_required(['admin'])
def analytics_dashboard():
    # Pass some initial stats if needed, or fetch all via API
    return render_template('admin/analytics.html')

@app.route('/analytics/data')
@login_required
@role_required(['admin'])
def analytics_data():
    time_range = request.args.get('range', '30d')
    now = datetime.now()
    
    if time_range == '7d':
        start_date = now - timedelta(days=7)
        group_by = 'day'
    elif time_range == '15d':
        start_date = now - timedelta(days=15)
        group_by = 'day'
    elif time_range == '6m':
        start_date = now - timedelta(days=180)
        group_by = 'month'
    elif time_range == '1y':
        start_date = now - timedelta(days=365)
        group_by = 'month'
    else:  # 30d default
        start_date = now - timedelta(days=30)
        group_by = 'day'

    def get_monthly_data(model, date_field, amount_field=None):
        group_expr = db.func.date(date_field) if group_by == 'day' else db.func.month(date_field)
        
        query = db.session.query(group_expr, db.func.count(model.id) if amount_field is None else db.func.sum(amount_field))
        query = query.filter(date_field >= start_date).group_by(group_expr)
        
        if model == User:
            query = query.filter(User.role == UserRole.STUDENT)
        if model == Fine:
            query = query.filter(Fine.paid == True)
            
        return query.order_by(group_expr).all()

    issues = get_monthly_data(Issue, Issue.issue_date)
    fees = get_monthly_data(FeeCollection, FeeCollection.collection_date, FeeCollection.total_amount)
    regs = get_monthly_data(User, User.created_at)
    fines = get_monthly_data(Fine, Fine.paid_date, Fine.amount)

    # Static data independent of time range
    cat_distribution = db.session.query(Category.name, db.func.count(Book.id)).join(Book).group_by(Category.name).all()
    top_borrowers = db.session.query(User.full_name, db.func.count(Issue.id)).join(Issue).filter(User.role == UserRole.STUDENT).group_by(User.full_name).order_by(db.func.count(Issue.id).desc()).limit(5).all()
    available_books = db.session.query(db.func.sum(Book.available_copies)).scalar() or 0
    total_copies = db.session.query(db.func.sum(Book.total_copies)).scalar() or 0
    issued_count = total_copies - available_books

    import calendar
    month_names = {i: name for i, name in enumerate(calendar.month_abbr) if i > 0}

    def format_labels(data):
        labels = []
        for x in data:
            if group_by == 'day':
                if hasattr(x[0], 'strftime'):
                    labels.append(x[0].strftime('%d %b'))
                else:
                    labels.append(str(x[0]))
            else:
                labels.append(month_names.get(x[0], f"M{x[0]}"))
        return labels

    return jsonify({
        'monthly_issues': {'labels': format_labels(issues), 'values': [x[1] for x in issues]},
        'monthly_fees': {'labels': format_labels(fees), 'values': [float(x[1] or 0) for x in fees]},
        'monthly_regs': {'labels': format_labels(regs), 'values': [x[1] for x in regs]},
        'monthly_fines': {'labels': format_labels(fines), 'values': [float(x[1] or 0) for x in fines]},
        'categories': {'labels': [x[0] for x in cat_distribution], 'values': [x[1] for x in cat_distribution]},
        'top_students': {'labels': [x[0] for x in top_borrowers], 'values': [x[1] for x in top_borrowers]},
        'book_status': {'labels': ['Available', 'Issued'], 'values': [int(available_books), int(issued_count)]}
    })

@app.route('/admin/export/analytics')
@login_required
@role_required(['admin'])
def export_analytics_excel():
    import pandas as pd
    import io
    from flask import send_file
    
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        # 1. Monthly Trends
        issues_by_month = db.session.query(
            db.func.month(Issue.issue_date), db.func.count(Issue.id)
        ).group_by(db.func.month(Issue.issue_date)).all()
        pd.DataFrame(issues_by_month, columns=['Month', 'Total Issues']).to_excel(writer, index=False, sheet_name='Monthly Trends')
        
        # 2. Book Categories
        cat_dist = db.session.query(Category.name, db.func.count(Book.id)).join(Book).group_by(Category.name).all()
        pd.DataFrame(cat_dist, columns=['Category', 'Book Count']).to_excel(writer, index=False, sheet_name='Inventory Data')
        
        # 3. Student Activity
        stu_act = db.session.query(User.full_name, User.username, db.func.count(Issue.id)).join(Issue).group_by(User.id).order_by(db.func.count(Issue.id).desc()).all()
        pd.DataFrame(stu_act, columns=['Student Name', 'Username', 'Total Issues']).to_excel(writer, index=False, sheet_name='Student Engagement')
        
        # 4. Financials
        fee_sum = db.session.query(StudentFee.status, db.func.sum(StudentFee.amount), db.func.count(StudentFee.id)).group_by(StudentFee.status).all()
        pd.DataFrame(fee_sum, columns=['Status', 'Total Amount', 'Count']).to_excel(writer, index=False, sheet_name='Financial Overview')
        
    output.seek(0)
    return send_file(output, 
                     download_name="Library_Intelligence_Report.xlsx", 
                     as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/admin/export/graphical_report')
@login_required
@role_required(['admin'])
def export_graphical_report():
    import pandas as pd
    import io
    from flask import send_file
    
    output = io.BytesIO()
    # Note: We use the xlsxwriter engine directly to manage workbook/charts
    writer = pd.ExcelWriter(output, engine='xlsxwriter')
    workbook = writer.book
    
    # 1. Monthly Trends with Chart
    issues_by_month = db.session.query(
        db.func.month(Issue.issue_date), db.func.count(Issue.id)
    ).group_by(db.func.month(Issue.issue_date)).all()
    df_month = pd.DataFrame(issues_by_month, columns=['Month', 'Issues'])
    
    sheet_name = 'Trends'
    df_month.to_excel(writer, sheet_name=sheet_name, index=False)
    ws_trends = writer.sheets[sheet_name]
        
    chart_trends = workbook.add_chart({'type': 'line'})
    chart_trends.add_series({
        'name':       'Monthly Issues',
        'categories': [sheet_name, 1, 0, len(df_month), 0],
        'values':     [sheet_name, 1, 1, len(df_month), 1],
        'marker':     {'type': 'circle', 'size': 8},
        'line':       {'color': '#4F46E5', 'width': 3.25},
    })
    chart_trends.set_title({'name': 'Library Borrowing Trends'})
    chart_trends.set_x_axis({'name': 'Month Index'})
    chart_trends.set_y_axis({'name': 'Total Issues'})
    ws_trends.insert_chart('D2', chart_trends)

    # 2. Categories with Pie Chart
    cat_dist = db.session.query(Category.name, db.func.count(Book.id)).join(Book).group_by(Category.name).all()
    df_cat = pd.DataFrame(cat_dist, columns=['Category', 'Count'])
    
    sheet_cat = 'Inventory'
    df_cat.to_excel(writer, sheet_name=sheet_cat, index=False)
    ws_cat = writer.sheets[sheet_cat]
        
    chart_cat = workbook.add_chart({'type': 'pie'})
    chart_cat.add_series({
        'name':       'Book Distribution',
        'categories': [sheet_cat, 1, 0, len(df_cat), 0],
        'values':     [sheet_cat, 1, 1, len(df_cat), 1],
        'data_labels': {'percentage': True},
    })
    chart_cat.set_title({'name': 'Books by Category'})
    ws_cat.insert_chart('E2', chart_cat)

    # 3. Financials with Column Chart
    fee_sum = db.session.query(StudentFee.status, db.func.count(StudentFee.id)).group_by(StudentFee.status).all()
    df_fee = pd.DataFrame(fee_sum, columns=['Status', 'Count'])
    
    sheet_fee = 'Financials'
    df_fee.to_excel(writer, sheet_name=sheet_fee, index=False)
    ws_fee = writer.sheets[sheet_fee]
        
    chart_fee = workbook.add_chart({'type': 'column'})
    chart_fee.add_series({
        'name':       'Item Count',
        'categories': [sheet_fee, 1, 0, len(df_fee), 0],
        'values':     [sheet_fee, 1, 1, len(df_fee), 1],
        'fill':       {'color': '#F59E0B'}
    })
    chart_fee.set_title({'name': 'Fee Status (Item Distribution)'})
    ws_fee.insert_chart('D2', chart_fee)
    
    writer.close()
    output.seek(0)
    return send_file(output, 
                     download_name="Library_Graphical_Insights.xlsx", 
                     as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# --- FINE MANAGEMENT ---

@app.route('/admin/fines')
@login_required
@role_required(['admin', 'librarian'])
def fine_list():
    page = request.args.get('page', 1, type=int)
    search = request.args.get('q', '')
    
    query = Fine.query.join(Issue).join(User)
    if search:
        query = query.filter(User.full_name.ilike(f'%{search}%') | User.username.ilike(f'%{search}%'))
        
    fines = query.order_by(Fine.paid.asc(), Fine.id.desc()).paginate(page=page, per_page=15)
    return render_template('admin/fines.html', fines=fines, search=search)

@app.route('/admin/fines/pay/<int:fine_id>', methods=['POST'])
@login_required
@role_required(['admin', 'librarian'])
def pay_fine(fine_id):
    fine = Fine.query.get_or_404(fine_id)
    if fine.paid:
        flash('Fine already paid.', 'info')
    else:
        fine.paid = True
        fine.paid_date = datetime.utcnow()
        db.session.commit()
        flash(f'Fine of {fine.amount} marked as paid.', 'success')
    return redirect(url_for('fine_list'))


# --- FEES MODULE ---

@app.route('/admin/fees/headers')
@login_required
@role_required(['admin'])
def fee_headers():
    headers = FeeHeader.query.order_by(FeeHeader.priority.desc()).all()
    return render_template('admin/fees/fee_headers.html', fee_headers=headers)

@app.route('/admin/fees/save_header', methods=['POST'])
@login_required
@role_required(['admin'])
def save_fee_header():
    try:
        header_id = request.form.get('header_id')
        name = request.form.get('name')
        amount = float(request.form.get('amount'))
        priority = int(request.form.get('priority'))
        admission_type = request.form.get('admission_type')
        applicable_for = request.form.get('applicable_for')
        status = True if request.form.get('status') else False

        due_date_str = request.form.get('due_date')
        end_date_str = request.form.get('end_date')

        due_date = datetime.strptime(due_date_str, '%Y-%m-%d') if due_date_str else None
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d') if end_date_str else None

        if header_id:
            # Update
            header = FeeHeader.query.get_or_404(int(header_id))
            header.name = name
            header.amount = amount
            header.priority = priority
            header.admission_type = admission_type
            header.applicable_for = applicable_for
            header.status = status
            header.due_date = due_date
            header.end_date = end_date
            flash('Fee Header updated successfully!', 'success')
        else:
            # Create
            new_header = FeeHeader(
                name=name, amount=amount, priority=priority,
                admission_type=admission_type, applicable_for=applicable_for,
                status=status, due_date=due_date, end_date=end_date
            )
            db.session.add(new_header)
            flash('Fee Header created successfully!', 'success')
        
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        flash(f'Error saving fee header: {str(e)}', 'danger')

    return redirect(url_for('fee_headers'))

@app.route('/api/search_users')
@login_required
def search_users():
    q = request.args.get('q', '').lower()
    role = request.args.get('role', 'student')
    
    if not q:
        return jsonify([])

    query = User.query.filter(User.role == UserRole.STUDENT if role == 'student' else True)
    query = query.filter(
        (User.full_name.ilike(f'%{q}%')) | 
        (User.username.ilike(f'%{q}%')) | 
        (User.enrollment_number.ilike(f'%{q}%'))
    )
    
    results = query.limit(10).all()
    data = []
    for u in results:
        data.append({
            'id': u.id,
            'full_name': u.full_name,
            'username': u.username,
            'enrollment_number': u.enrollment_number,
            'program': u.program,
            'semester': u.semester,
            'contact_number': u.contact_number
        })
    return jsonify(data)

@app.route('/admin/fees/apply')
@login_required
@role_required(['admin'])
def apply_fees():
    headers = FeeHeader.query.filter_by(status=True).order_by(FeeHeader.priority.desc()).all()
    return render_template('admin/fees/apply_fees.html', fee_headers=headers)

@app.route('/admin/fees/save_student_fees', methods=['POST'])
@login_required
@role_required(['admin'])
def save_student_fees():
    student_id = request.form.get('student_id')
    fee_ids = request.form.getlist('fee_ids')
    
    if not student_id or not fee_ids:
        flash('Please select a student and at least one fee.', 'warning')
        return redirect(url_for('apply_fees'))

    try:
        count = 0
        for fid in fee_ids:
            # Get overrides
            amount_str = request.form.get(f'amount_{fid}')
            due_date_str = request.form.get(f'due_date_{fid}')
            
            # Fetch original header for defaults if needed (though form provides values)
            # Trusted form values for now as "Overrides"
            amount = float(amount_str) if amount_str else 0.0
            due_date = datetime.strptime(due_date_str, '%Y-%m-%d') if due_date_str else None
            
            # Check duplication? User requirement: "Fees applied once should not duplicate on re-apply"
            # We can check if pending fee of same header exists for student
            existing = StudentFee.query.filter_by(
                student_id=student_id, 
                fee_header_id=fid, 
                status='Pending'
            ).first()
            
            if existing:
                continue # Skip duplicate
            
            new_fee = StudentFee(
                student_id=student_id,
                fee_header_id=fid,
                amount=amount,
                due_date=due_date,
                status='Pending'
            )
            db.session.add(new_fee)
            count += 1
        db.session.commit()
        if count > 0:
            flash(f'{count} fees applied successfully!', 'success')
        else:
            flash('Selected fees already applied to this student.', 'info')
    except Exception as e:
        db.session.rollback()
        flash(f'Error applying fees: {str(e)}', 'danger')

    return redirect(url_for('apply_fees'))

@app.route('/admin/fees/bulk_template')
@login_required
@role_required(['admin'])
def download_fee_bulk_template():
    import pandas as pd
    import io
    from flask import send_file

    # Get all active students
    students = User.query.filter_by(role=UserRole.STUDENT, is_active=True).all()
    # Get all active fee headers
    headers = FeeHeader.query.filter_by(status=True).order_by(FeeHeader.priority.desc()).all()
    
    # Headers for Excel
    cols = ['Username', 'Full Name', 'Enrollment', 'Program', 'Semester']
    for h in headers:
        cols.append(h.name)
        
    data = []
    for s in students:
        row = {
            'Username': s.username,
            'Full Name': s.full_name,
            'Enrollment': s.enrollment_number,
            'Program': s.program,
            'Semester': s.semester
        }
        for h in headers:
            # Check if this fee is already applied to student to preserve current state or just put defaults
            row[h.name] = h.amount # Default value from header
        data.append(row)
        
    df = pd.DataFrame(data, columns=cols)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False, sheet_name='BulkFees')
    
    output.seek(0)
    return send_file(output, 
                     download_name="Bulk_Fees_Apply_Template.xlsx", 
                     as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/admin/fees/bulk_import', methods=['GET', 'POST'])
@login_required
@role_required(['admin'])
def bulk_import_fees():
    if request.method == 'POST':
        import pandas as pd
        file = request.files.get('file')
        if not file or not file.filename.endswith(('.xlsx', '.xls')):
            flash('Please upload a valid Excel file.', 'danger')
            return redirect(url_for('bulk_import_fees'))
            
        try:
            df = pd.read_excel(file)
            # Find fee headers mapping from columns
            headers = FeeHeader.query.filter_by(status=True).all()
            header_map = {h.name: h.id for h in headers}
            
            applied_count = 0
            for _, row in df.iterrows():
                username = str(row.get('Username', '')).strip()
                if username == 'nan' or not username: continue
                
                student = User.query.filter_by(username=username, role=UserRole.STUDENT).first()
                if not student: continue
                
                for col_name in df.columns:
                    if col_name in header_map:
                        amount_val = row.get(col_name)
                        if pd.isna(amount_val): continue
                        
                        try:
                            amount = float(amount_val)
                        except:
                            continue

                        if amount > 0:
                            hid = header_map[col_name]
                            # Check if already applied
                            existing = StudentFee.query.filter_by(student_id=student.id, fee_header_id=hid, status='Pending').first()
                            if not existing:
                                h_obj = FeeHeader.query.get(hid)
                                new_fee = StudentFee(
                                    student_id=student.id,
                                    fee_header_id=hid,
                                    amount=amount,
                                    due_date=h_obj.due_date,
                                    status='Pending'
                                )
                                db.session.add(new_fee)
                                applied_count += 1
                                
            db.session.commit()
            flash(f'Successfully processed bulk fees. {applied_count} fee items applied!', 'success')
            return redirect(url_for('apply_fees'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error importing fees: {str(e)}', 'danger')
            
    return render_template('admin/fees/bulk_fees_page.html')

@app.route('/admin/fees/collection')
@login_required
@role_required(['admin'])
def fee_collection():
    return render_template('admin/fees/collection.html', now=datetime.utcnow())

@app.route('/api/student_fee_data/<int:student_id>')
@login_required
def get_student_fee_data(student_id):
    student = User.query.get_or_404(student_id)
    
    # Get pending fees
    pending_fees = StudentFee.query.filter_by(student_id=student_id).all()
    fees_list = []
    for f in pending_fees:
        # Calculate already collected for this specific item
        collected = sum(item.amount_collected for item in f.collection_entries)
        payable = f.amount - collected
        
        if payable > 0 or f.status == 'Pending':
            fees_list.append({
                'id': f.id,
                'name': f.fee_header.name,
                'due_date': f.due_date.strftime('%Y-%m-%d') if f.due_date else 'N/A',
                'amount': f.amount,
                'collected': collected,
                'payable': payable,
                'status': f.status
            })
            
    # Get previous collections
    collections = FeeCollection.query.filter_by(student_id=student_id).order_by(FeeCollection.collection_date.desc()).all()
    history = []
    for c in collections:
        # Get fee types for this collection
        fee_types = list(set([item.student_fee.fee_header.name for item in c.items]))
        history.append({
            'id': c.id,
            'voucher_no': c.voucher_no,
            'date': c.collection_date.strftime('%Y-%m-%d %H:%M'),
            'acad_year': c.academic_year,
            'fee_type': ", ".join(fee_types) if fee_types else 'N/A',
            'payment_mode': c.payment_mode,
            'ref_no': c.transaction_no or '-',
            'amount': c.total_amount,
            'late_fees': c.late_fees,
            'addl_charges': c.additional_charges,
            'remarks': c.remarks or '-',
            # Student fields for table (requested in ERP)
            'stu_name': student.full_name,
            'roll_no': student.username,
            'enroll': student.enrollment_number or '-',
            'contact': student.contact_number or '-'
        })
        
    return jsonify({
        'pending_fees': fees_list,
        'history': history,
        'student': {
            'id': student.id,
            'full_name': student.full_name,
            'username': student.username,
            'enrollment_number': student.enrollment_number,
            'program': student.program,
            'semester': student.semester,
            'contact_number': student.contact_number
        }
    })

@app.route('/admin/fees/receipt/<int:collection_id>')
@login_required
@role_required(['admin'])
def view_receipt(collection_id):
    collection = FeeCollection.query.get_or_404(collection_id)
    return render_template('admin/fees/receipt.html', collection=collection)

@app.route('/admin/fees/collect_payment', methods=['POST'])
@login_required
@role_required(['admin'])
def collect_payment():
    data = request.json
    student_id = data.get('student_id')
    payment_mode = data.get('payment_mode')
    receipt_date_str = data.get('receipt_date')
    total_paid = float(data.get('amount_paid', 0))
    late_fees = float(data.get('late_fees', 0))
    additional_charges = float(data.get('additional_charges', 0))
    remarks = data.get('remarks')
    
    fee_items = data.get('fee_items', []) # List of {id, collected, discount}

    try:
        # Generate Voucher Number
        year = datetime.utcnow().year
        last_coll = FeeCollection.query.order_by(FeeCollection.id.desc()).first()
        next_id = (last_coll.id + 1) if last_coll else 1
        voucher_no = f"VCH-{year}-{next_id:04d}"
        
        receipt_date = datetime.strptime(receipt_date_str, '%Y-%m-%d') if receipt_date_str else datetime.utcnow()

        collection = FeeCollection(
            voucher_no=voucher_no,
            collection_date=receipt_date,
            academic_year=data.get('academic_year', f"{year}-{year+1}"),
            payment_mode=payment_mode,
            total_amount=total_paid,
            late_fees=late_fees,
            additional_charges=additional_charges,
            bank_name=data.get('bank_name'),
            transaction_no=data.get('transaction_no'),
            transaction_date=datetime.strptime(data.get('transaction_date'), '%Y-%m-%d').date() if data.get('transaction_date') else None,
            remarks=remarks,
            student_id=student_id
        )
        db.session.add(collection)
        db.session.flush() # Get collection ID
        
        for item in fee_items:
            fee_id = item.get('id')
            amt = float(item.get('collected', 0))
            disc = float(item.get('discount', 0))
            
            if amt > 0 or disc > 0:
                coll_item = FeeCollectionItem(
                    collection_id=collection.id,
                    student_fee_id=fee_id,
                    amount_collected=amt,
                    discount=disc
                )
                db.session.add(coll_item)
                
                # Update StudentFee status
                s_fee = StudentFee.query.get(fee_id)
                current_coll = sum(e.amount_collected for e in s_fee.collection_entries)
                if current_coll >= s_fee.amount:
                    s_fee.status = 'Paid'
                elif current_coll > 0:
                    s_fee.status = 'Partial'

        db.session.commit()
        return jsonify({'success': True, 'voucher_no': voucher_no, 'collection_id': collection.id})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)})

# --- BOOK REQUESTS & NOTIFICATIONS ---

@app.route('/student/request_book/<int:book_id>', methods=['POST'])
@login_required
@role_required(['student'])
def request_book(book_id):
    book = Book.query.get_or_404(book_id)
    if book.available_copies < 1:
        flash('This book is currently out of stock.', 'warning')
        return redirect(url_for('book_list'))
    
    # Check if already requested
    existing = BookRequest.query.filter_by(user_id=current_user.id, book_id=book_id, status='pending').first()
    if existing:
        flash('You already have a pending request for this book.', 'info')
        return redirect(url_for('book_list'))

    new_request = BookRequest(user_id=current_user.id, book_id=book_id)
    db.session.add(new_request)
    
    # Create notifications for all admins and librarians
    managers = User.query.filter(User.role.in_([UserRole.ADMIN, UserRole.LIBRARIAN])).all()
    for mgr in managers:
        notif = Notification(
            user_id=mgr.id,
            message=f"New book request from {current_user.full_name} for '{book.title}'",
            link=url_for('view_book_requests')
        )
        db.session.add(notif)
    
    db.session.commit()
    flash('Book request sent to librarian!', 'success')
    return redirect(url_for('book_list'))

@app.route('/admin/book_requests')
@login_required
@role_required(['admin', 'librarian'])
def view_book_requests():
    requests = BookRequest.query.filter_by(status='pending').order_by(BookRequest.request_date.desc()).all()
    return render_template('admin/book_requests.html', book_requests=requests)

@app.route('/admin/handle_request/<int:request_id>/<action>', methods=['POST'])
@login_required
@role_required(['admin', 'librarian'])
def handle_book_request(request_id, action):
    req = BookRequest.query.get_or_404(request_id)
    if action == 'approve':
        if req.book_obj.available_copies < 1:
            flash('Book no longer available.', 'danger')
            return redirect(url_for('view_book_requests'))
        
        # Create issue
        new_issue = Issue(
            user_id=req.user_id,
            book_id=req.book_id,
            due_date=datetime.utcnow() + timedelta(days=14),
            status='issued'
        )
        req.book_obj.available_copies -= 1
        req.status = 'approved'
        db.session.add(new_issue)
        
        # Notify student
        notif = Notification(
            user_id=req.user_id,
            message=f"Your request for '{req.book_obj.title}' has been APPROVED!",
            link=url_for('dashboard')
        )
        db.session.add(notif)
        flash('Request approved and book issued.', 'success')
        
    elif action == 'reject':
        req.status = 'rejected'
        # Notify student
        notif = Notification(
            user_id=req.user_id,
            message=f"Your request for '{req.book_obj.title}' was rejected.",
            link=url_for('book_list')
        )
        db.session.add(notif)
        flash('Request rejected.', 'info')
        
    db.session.commit()
    return redirect(url_for('view_book_requests'))

@app.route('/api/notifications')
@login_required
def get_notifications():
    notifs = Notification.query.filter_by(user_id=current_user.id).order_by(Notification.timestamp.desc()).limit(10).all()
    unread_count = Notification.query.filter_by(user_id=current_user.id, is_read=False).count()
    return jsonify({
        'notifications': [{
            'id': n.id,
            'message': n.message,
            'link': n.link,
            'is_read': n.is_read,
            'time': n.timestamp.strftime('%H:%M %p')
        } for n in notifs],
        'unread_count': unread_count
    })

@app.route('/admin/fees/bulk_collection_template')
@login_required
@role_required(['admin'])
def download_bulk_collection_template():
    import pandas as pd
    import io
    from flask import send_file
    
    # Get students with at least one pending fee
    pending_students = db.session.query(User).join(StudentFee).filter(StudentFee.status.in_(['Pending', 'Partial'])).distinct().all()
    
    cols = ['Username', 'Full Name', 'Enrollment', 'Program', 'Semester', 'Total Due', 
            'Payable Amount', 'Discount', 'Payment Mode', 'Receipt Date', 
            'Late Fees', 'Additional Charges', 'Ref ID / Trans No', 'Remarks']
    
    data = []
    for s in pending_students:
        # Calculate total due for this student across all pending items
        pending_items = StudentFee.query.filter_by(student_id=s.id).filter(StudentFee.status.in_(['Pending', 'Partial'])).all()
        total_due = 0
        for f in pending_items:
            # Collected so far for this specific item
            collected = sum(item.amount_collected for item in f.collection_entries)
            total_due += (f.amount - collected)
            
        if total_due > 0:
            data.append({
                'Username': s.username,
                'Full Name': s.full_name,
                'Enrollment': s.enrollment_number,
                'Program': s.program,
                'Semester': s.semester,
                'Total Due': total_due,
                'Payable Amount': total_due, # Suggesting full payment
                'Discount': 0,
                'Payment Mode': 'Cash',
                'Receipt Date': datetime.utcnow().strftime('%Y-%m-%d'),
                'Late Fees': 0,
                'Additional Charges': 0,
                'Ref ID / Trans No': '',
                'Remarks': 'Bulk Import'
            })
            
    df = pd.DataFrame(data, columns=cols)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False, sheet_name='FeeCollection')
        
    output.seek(0)
    return send_file(output, 
                     download_name="Bulk_Collection_Template.xlsx", 
                     as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/admin/fees/bulk_collection_import', methods=['GET', 'POST'])
@login_required
@role_required(['admin'])
def bulk_collection_import():
    if request.method == 'POST':
        import pandas as pd
        file = request.files.get('file')
        if not file or not file.filename.endswith(('.xlsx', '.xls')):
            flash('Please upload a valid Excel file.', 'danger')
            return redirect(url_for('bulk_collection_import'))
            
        try:
            df = pd.read_excel(file)
            import_count = 0
            
            for _, row in df.iterrows():
                username = str(row.get('Username', '')).strip()
                if username == 'nan' or not username: continue
                
                student = User.query.filter_by(username=username, role=UserRole.STUDENT).first()
                if not student: continue
                
                amount_paid = float(row.get('Payable Amount', 0))
                if amount_paid <= 0: continue
                
                discount = float(row.get('Discount', 0))
                late_fees = float(row.get('Late Fees', 0))
                add_charges = float(row.get('Additional Charges', 0))
                mode = str(row.get('Payment Mode', 'Cash'))
                date_str = str(row.get('Receipt Date', datetime.utcnow().strftime('%Y-%m-%d')))
                trans_no = str(row.get('Ref ID / Trans No', ''))
                remarks = str(row.get('Remarks', 'Bulk Collection'))
                
                # Logic copied from collect_payment to maintain consistency
                # 1. Generate Voucher
                year = datetime.utcnow().year
                last_coll = FeeCollection.query.order_by(FeeCollection.id.desc()).first()
                next_id = (last_coll.id + 1) if last_coll else 1
                voucher_no = f"VCH-{year}-{next_id:04d}"
                
                try:
                    receipt_date = datetime.strptime(date_str[:10], '%Y-%m-%d')
                except:
                    receipt_date = datetime.utcnow()

                collection = FeeCollection(
                    voucher_no=voucher_no,
                    collection_date=receipt_date,
                    academic_year=f"{year}-{year+1}",
                    payment_mode=mode,
                    total_amount=amount_paid + late_fees + add_charges,
                    late_fees=late_fees,
                    additional_charges=add_charges,
                    transaction_no=trans_no if trans_no != 'nan' else '',
                    remarks=remarks if remarks != 'nan' else '',
                    student_id=student.id
                )
                db.session.add(collection)
                db.session.flush()

                # 2. Distribute amount across pending fees FIFO (First In First Out)
                pending_items = StudentFee.query.filter_by(student_id=student.id).filter(StudentFee.status.in_(['Pending', 'Partial'])).order_by(StudentFee.due_date.asc()).all()
                
                remaining_to_distribute = amount_paid
                for f in pending_items:
                    if remaining_to_distribute <= 0: break
                    
                    collected_so_far = sum(item.amount_collected for item in f.collection_entries)
                    balance = f.amount - collected_so_far
                    
                    if balance <= 0: continue
                    
                    to_apply = min(remaining_to_distribute, balance)
                    
                    coll_item = FeeCollectionItem(
                        collection_id=collection.id,
                        student_fee_id=f.id,
                        amount_collected=to_apply,
                        discount=discount if remaining_to_distribute == amount_paid else 0 # Apply discount only once or split? Simple: only once.
                    )
                    db.session.add(coll_item)
                    
                    remaining_to_distribute -= to_apply
                    
                    # Update status
                    if (collected_so_far + to_apply) >= f.amount:
                        f.status = 'Paid'
                    else:
                        f.status = 'Partial'

                import_count += 1
                
            db.session.commit()
            flash(f'Successfully processed bulk collection. {import_count} student payments recorded!', 'success')
            return redirect(url_for('fee_collection'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error importing collection: {str(e)}', 'danger')
            
    return render_template('admin/fees/bulk_collection_page.html')

@app.route('/api/notifications/mark_read', methods=['POST'])
@login_required
def mark_notifications_read():
    Notification.query.filter_by(user_id=current_user.id, is_read=False).update({Notification.is_read: True})
    db.session.commit()
    return jsonify({'success': True})

if __name__ == '__main__':
    app.run(debug=True)

